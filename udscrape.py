#Main file
#This file creates the GUI for the Udemy Coupon Scraper
#Version 0: By MikeePy

import webgetter
import openpyxl as px
from tkinter import *
scrape_list = []
flag = 0
def scrapecommand():
    #initiate scrape for udemy coupons
    list.delete(0,END)
    list.insert(END,"Please Wait, Scraping is under way...")
    scrbox = page_text.get()
    global flag
    global scrape_list
    try:
        page_num = int(scrbox)
        
        scrape_list = webgetter.getlist(page_num)
        flag = 1
        list.delete(0,END)
        for row in scrape_list:
            list.insert(END,row)
    except ValueError:
        list.insert(END,"Wrong input, input valid number")
    
    
def excelcommand():
    #Outputs to excel spreadsheet on same directory
    if flag == 0:
        list.delete(0,END)
        list.insert(END,"Please scrape first")
    else:
        print(scrape_list)
        workbook = px.Workbook()
        wsheet = workbook.active
        wsheet.cell(row=1, column = 1).value = "Title"
        wsheet.cell(row=1, column = 2).value = "Link"
        count = 2
        for line in scrape_list:
            title = line[0]
            link = line[1]
            wsheet.cell(row=count, column=1).value = title
            wsheet.cell(row=count, column=2).value = link
            count = count+1       
        workbook.save('udemy coupon output.xlsx')
        print("File Created: udemy coupon output.xlsx")

win = Tk()

win.wm_title('Udemy Coupon Scraper')
l1 = Label(win, text='Number of Pages to Scrape')
l1.grid(row=0, column=0)

page_text = StringVar()
e1 = Entry(win, textvariable = page_text)
e1.grid(row=1,column=0)
b1 = Button(win, text='Scrape', width = 12, pady=5, command=scrapecommand)
b1.grid(row=0,column=2)
b2 = Button(win, text='Export to Excel', width = 12, pady=5, command=excelcommand)
b2.grid(row=1,column=2)
list = Listbox(win, height=8, width=75)

list.grid(row=2, column = 0, rowspan = 9, columnspan = 6)


win.mainloop()